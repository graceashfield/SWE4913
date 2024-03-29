import sys
import os
import copy
from PyQt6 import QtGui
from PyQt6.QtCore import Qt, QTimer, QCoreApplication
import numpy as np
import pandas as pd
import openpyxl
from PyQt6.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QGroupBox, QLabel, QGridLayout, \
    QMainWindow, QWidget, QPushButton, QFileDialog, QMessageBox, QFrame, QAbstractItemView, QLineEdit, QComboBox, \
    QProgressDialog
from openpyxl.styles import Font


def paint_gradient_bar(widget, event):
    """
        Paints a gradient bar on the given widget.
        :param widget: The widget to paint on
        :type widget: QtGui.QWidget
        :param event: The paint event
        :type event: QtGui.QPaintEvent
        """
    painter = QtGui.QPainter(widget)
    # gradient = QtGui.QLinearGradient(0, 0, 0, widget.height())
    gradient = QtGui.QLinearGradient(0, 0, widget.width(), 0)
    gradient.setColorAt(0, QtGui.QColor(254, 0, 2))
    gradient.setColorAt(0.2, QtGui.QColor(216, 0, 39))
    gradient.setColorAt(0.4, QtGui.QColor(161, 1, 93))
    gradient.setColorAt(0.6, QtGui.QColor(99, 0, 158))
    gradient.setColorAt(0.8, QtGui.QColor(68, 0, 204))
    gradient.setColorAt(1, QtGui.QColor(3, 2, 252))
    painter.fillRect(event.rect(), gradient)
    painter.end()


def create_legend():
    """
        Creates a legend widget with a gradient bar and labels.
        :return: The legend widget
        :rtype: QtGui.QWidget
        """
    legend_widget = QWidget()
    layout = QGridLayout()
    legend_widget.setLayout(layout)
    layout.addWidget(QLabel("Transition Frames"), 0, 0, Qt.AlignmentFlag.AlignCenter)
    layout.addWidget(QLabel("More Concentrated"), 0, 1, Qt.AlignmentFlag.AlignLeft)
    layout.addWidget(QLabel("Less Concentrated"), 0, 3, Qt.AlignmentFlag.AlignRight)
    layout.setRowMinimumHeight(0, 10)
    layout.setRowMinimumHeight(1, 20)
    gradient_bar = QFrame()
    gradient_bar.setFrameShape(QFrame.Shape.Box)
    gradient_bar.setMaximumSize(410, 40)
    gradient_bar.paintEvent = lambda e: paint_gradient_bar(gradient_bar, e)
    layout.addWidget(gradient_bar, 1, 1, 1, 3)
    return legend_widget


def get_colour(value):
    """
        Returns a QColor object based on the given value.
        :param value: The value used to determine the color
        :type value: int
        :return: The QColor object
        :rtype: QtGui.QColor
        """
    colour = QtGui.QColor(255, 255, 255)
    if value == 0:
        colour = QtGui.QColor(3, 2, 252)
    elif 1 <= value <= 10:
        colour = QtGui.QColor(0, 0, 175)
    elif 10 < value <= 20:
        colour = QtGui.QColor(68, 0, 204)
    elif 20 < value <= 30:
        colour = QtGui.QColor(70, 21, 169)
    elif 30 < value <= 40:
        colour = QtGui.QColor(99, 0, 158)
    elif 40 < value <= 50:
        colour = QtGui.QColor(102, 0, 102)
    elif 50 < value <= 60:
        colour = QtGui.QColor(161, 1, 93)
    elif 60 < value <= 70:
        colour = QtGui.QColor(153, 0, 0)
    elif 70 < value <= 80:
        colour = QtGui.QColor(216, 0, 39)
    elif 80 < value <= 90:
        colour = QtGui.QColor(204, 0, 0)
    elif 90 < value:
        colour = QtGui.QColor(254, 0, 2)
    return colour


def sort_key(item):
    """
        Returns a tuple used for sorting a list of items.
        The function extracts the first & second elements from the first element of the input item and returns a tuple.
        :param item: The item to extract the elements from
        :type item: tuple
        :return: A tuple used for sorting the items
        :rtype: tuple
        """
    return item[0][0], item[0][1]


def count_percentage(numbers):
    """
        Calculates the percentage of occurrences of each number in the input list.
        :param numbers: A list of integers
        :type numbers: list
        :return: A dictionary where the keys are the numbers and the values are the percentage of occurrences
        :rtype: dict
        """
    count = {}
    total = len(numbers)
    for num in numbers:
        if num not in count:
            count[num] = 0
        count[num] += 1
    # if a classifier did not appear during transition, count as 0
    for num in range(1, 8):
        if num not in count:
            count[num] = 0
    for num, num_count in count.items():
        count[num] = round((num_count / total) * 100)
    return count


def clear_table(table, row_num):
    """
        Clear the contents of each cell in the given row of the table, excluding headers.
        :param table: The QTableWidget object to clear
        :type table: QTableWidget
        :param row_num: The row number to clear
        :type row_num: int
    """
    for row in range(table.rowCount()+1):
        for col in range(table.columnCount()):
            table.item(row_num, col).setBackground(QtGui.QColor(220, 220, 220))
            item = table.item(row, col)
            if item is not None:
                item.setText("")
                item.setBackground(QtGui.QColor("transparent"))


class Classifier(QMainWindow):
    def __init__(self):
        super().__init__()

        self.file_names = None
        self.file_name = None
        self.cur_individual = 0
        self.cur_data = {}
        self.final_dict = {}
        self.all_data = {}
        self.dictio = {1: "NM", 2: "WF", 3: "WE", 4: "WP", 5: "WS", 6: "CG", 7: "HO"}

        self.setWindowTitle("EMG Classifier")
        self.showMaximized()
        layout = QGridLayout()

        # source file select
        groupbox0 = QGroupBox()
        grid0 = QGridLayout()
        self.data_set = QLabel("Please Select A Dataset")
        file_button = QPushButton("Select Data")
        file_button.clicked.connect(self.file_select)
        grid0.addWidget(self.data_set, 0, 0, Qt.AlignmentFlag.AlignCenter)
        grid0.addWidget(file_button, 1, 0)
        groupbox0.setLayout(grid0)
        groupbox0.setMaximumSize(200, 100)
        layout.addWidget(groupbox0, 0, 0)

        # UI layout
        groupbox1 = QGroupBox()
        grid1 = QGridLayout()
        groupbox1.setLayout(grid1)
        self.data_button = QPushButton("View Individual Data")
        self.data_button.clicked.connect(self.data_button_clicked)
        self.data_button.setCheckable(True)
        self.data_button.setEnabled(False)

        self.prev_button = QPushButton("Previous Data")
        self.next_button = QPushButton("Next Data")
        self.prev_button.setFixedWidth(130)
        self.next_button.setFixedWidth(130)
        self.prev_button.clicked.connect(self.prev_next)
        self.next_button.clicked.connect(self.prev_next)
        self.prev_button.hide()
        self.next_button.hide()
        self.file_selector = QComboBox()
        self.file_selector.currentIndexChanged.connect(self.data_update)
        self.file_selector.hide()
        self.current_field = QLabel("")
        self.current_field.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.current_field.hide()
        self.input_box = QLineEdit()  # Create a QLineEdit object
        self.input_box.setFixedWidth(100)  # Set maximum width
        self.input_box.timer = QTimer()  # timer triggers after 3 seconds of inactivity
        self.input_box.timer.setInterval(3000)  # 3000ms = 3s
        self.input_box.timer.setSingleShot(True)
        # Connect the signals and slots
        self.input_box.returnPressed.connect(self.data_update)

        self.input_box.hide()
        grid1.addWidget(self.prev_button, 2, 1)
        grid1.addWidget(self.next_button, 2, 3)
        grid1.addWidget(self.data_button, 0, 1)
        grid1.addWidget(self.current_field, 0, 2)
        grid1.addWidget(self.file_selector, 0, 3)
        grid1.addWidget(self.input_box, 2, 2)
        layout.addWidget(groupbox1, 0, 1, 1, 2)

        groupbox2 = QGroupBox()
        grid2 = QGridLayout()
        groupbox2.setLayout(grid2)
        self.row_button = QPushButton("View All Rows")
        self.row_button.clicked.connect(self.row_button_clicked)
        self.row_button.setCheckable(True)
        grid2.addWidget(self.row_button, 0, 3, 1, 1)
        groupbox2.setMaximumSize(200, 100)
        layout.addWidget(groupbox2, 0, 3)

        groupbox3 = QGroupBox()
        grid3 = QGridLayout()
        grid3.addWidget(create_legend(), 0, 0)
        groupbox3.setLayout(grid3)
        groupbox3.setMaximumSize(700, 100)
        layout.addWidget(groupbox3, 0, 4, 1, 2)

        # output
        groupbox4 = QGroupBox()
        self.grid4 = QGridLayout()
        self.all_tables = ["t1", "t2", "t3", "t4", "t5", "t6", "t7"]
        for x in range(len(self.all_tables)):
            cell_size = 45
            height = int(4.3 * cell_size)
            width = int(9.8 * cell_size)
            self.all_tables[x] = self.create_table(width, height, cell_size, x + 1)
            self.all_tables[x].hideRow(x)
            for col in range(self.all_tables[x].columnCount()):
                self.all_tables[x].item(x, col).setBackground(QtGui.QColor(220, 220, 220))

        self.grid4.addWidget(QLabel(), 1, 0)
        self.grid4.addWidget(self.all_tables[0], 0, 0, 1, 1)
        self.grid4.addWidget(self.all_tables[1], 0, 1, 1, 1)
        self.grid4.addWidget(self.all_tables[2], 0, 2, 1, 1)
        self.grid4.addWidget(self.all_tables[3], 1, 0, 1, 1)
        self.grid4.addWidget(self.all_tables[4], 1, 1, 1, 1)
        self.grid4.addWidget(self.all_tables[5], 1, 2, 1, 1)
        self.grid4.addWidget(self.all_tables[6], 2, 1, 1, 1)

        screenshot_button = QPushButton("Take Screenshot")
        screenshot_button.clicked.connect(self.screenshot)
        self.grid4.addWidget(screenshot_button, 2, 2, 1, 1)

        self.export_button = QPushButton('Export to Excel')
        self.export_button.setEnabled(False)
        self.export_button.clicked.connect(lambda: self.export_all_tables())
        self.grid4.addWidget(self.export_button, 2, 0, 1, 1)

        groupbox4.setLayout(self.grid4)
        layout.addWidget(groupbox4, 1, 0, 6, 6)

        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def data_update(self):
        """
            Updates the displayed data based on the user's input.
            Retrieves data from the file specified by the user's input, clears all tables, & updates fields
        """
        sender_button = app.sender()
        if sender_button == self.input_box:
            data_input = self.input_box.text().upper() + ".npz"
            if data_input in self.all_data:
                index = self.file_selector.findText(data_input)
                self.file_selector.setCurrentIndex(index)
            else:
                QMessageBox.critical(self, "Error", "Please input a valid file")
        else:
            data_input = self.file_selector.currentText()

        if data_input in self.all_data:
            self.cur_individual = list(self.all_data.keys()).index(data_input)
            for x in range(len(self.all_tables)):
                clear_table(self.all_tables[x], x)
            self.current_field.setText(str(data_input))
            self.percentage_data(self.all_data[data_input]["individual"])
            self.prev_button.setEnabled(self.cur_individual > 0)
            self.next_button.setEnabled(self.cur_individual < len(self.all_data.keys()) - 1)

    def file_select(self):
        """
            Opens a file dialog to select a directory containing NPZ files, then loads the data from the selected file.
            If the file contains the required data arrays, populates the UI with the data and enables the data button.
            If the file is not valid, displays an error message.
        """
        dir_path = str(QFileDialog.getExistingDirectory(self, "Select Directory"))
        if dir_path:
            self.clear()
            self.data_button.setChecked(False)
            self.data_button_clicked()
            self.file_names = os.listdir(dir_path)
            self.file_names.sort()
            self.file_selector.addItems(self.file_names)
            self.all_data = {}  # create a dictionary to hold all data for all files

            # Loading files
            for file_name in self.file_names:
                if os.path.isfile(os.path.join(dir_path, file_name)):
                    self.data_set.setText(os.path.basename(os.path.normpath(dir_path)))
                    file_path = os.path.join(dir_path, file_name)
                    try:
                        with np.load(file_path, allow_pickle=False) as fd:  # read data files
                            data = {"predictions": fd["predictions"], "probabilities": fd["probabilities"],
                                    "trials": fd["trials"], "truth": fd["ground_truth_table"],
                                    "steady": fd["steady_state_table"],
                                    "individual": {}}  # create a dictionary to hold data for each file
                            self.all_data[file_name] = data  # add data for this file to the all_data dictionary
                    except KeyError:
                        QMessageBox.critical(self, "Error", "Please select a valid file")
                        break
                    except ValueError:
                        QMessageBox.critical(self, "Error", "Cannot load file containing invalid data")
                        break
                else:
                    QMessageBox.critical(self, "Error", "Please select a valid directory")
                    break

            # Call self.data() once after all files have been read and their data has been stored
            self.data(self.all_data)
            self.data_button.setEnabled(True)
            self.export_button.setEnabled(True)

    def create_table(self, width, height, cell, num):
        """
            Creates a QTableWidget instance with customized properties and headers.
            :param width: the maximum width of the table
            :type width: int
            :param height: the maximum height of the table
            :type height: int
            :param cell: the size of each cell (width and height are proportional to it)
            :type cell: int
            :param num: the index of the vertical header to be removed
            :type num: int
            :return: a QTableWidget instance with customized properties and headers
            :rtype: QTableWidget
        """
        table = QTableWidget()
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        h_headers = self.dictio.copy()
        v_headers = h_headers.pop(num)
        header_labels = list(self.dictio.values())
        table.setRowCount(len(header_labels))
        header_labels.append("Frames")
        table.setColumnCount(len(header_labels))
        table.setHorizontalHeaderLabels(header_labels)

        count = 0
        for value in header_labels:
            item = QTableWidgetItem(v_headers + "->" + value)
            table.setVerticalHeaderItem(count, item)
            count += 1

        table.setMaximumSize(width, height)
        for x in range(len(header_labels)):
            table.setColumnWidth(x, cell)
            if x == 7:
                table.setColumnWidth(x, cell + 20)
            table.setRowHeight(x, round(cell / 1.6))
            for y in range(len(header_labels)):
                table.setItem(y, x, QTableWidgetItem(""))
        return table

    def row_button_clicked(self):
        """
            Function to handle the "View All Rows" / "View Less Rows" button clicked event.
            When button is checked, show all rows in the tables and increases their maximum size.
            When button is unchecked, hide all rows in the tables except for the headers & reduces their maximum size
        """
        if self.row_button.isChecked():
            self.row_button.setText("View Less Rows")
            for x in range(len(self.all_tables)):
                self.all_tables[x].showRow(x)
                self.all_tables[x].setMaximumSize(441, 220)
        else:  # if it is unchecked
            self.row_button.setText("View All Rows")
            for x in range(len(self.all_tables)):
                self.all_tables[x].hideRow(x)
                self.all_tables[x].setMaximumSize(441, 194)

    def data_button_clicked(self):
        """
            This function is called when the "View Individual Data" toggle button is clicked.
            If the button is checked, it displays the data for the currently selected field.
            If the button is unchecked, it returns to displaying the overall data.
        """
        if self.data_button.isChecked():
            self.cur_individual = 0
            self.data_button.setText("View All Data")
            for x in range(len(self.all_tables)):
                clear_table(self.all_tables[x], x)
            data = list(self.all_data.keys())[self.cur_individual]
            self.current_field.setText(str(data))
            self.percentage_data(self.all_data[data]["individual"])
            self.current_field.show()
            self.file_selector.show()
            self.next_button.show()
            self.prev_button.show()
            self.input_box.show()
        else:  # if it is unchecked
            self.cur_individual = 0
            self.data_button.setText("View Individual Data")
            for x in range(len(self.all_tables)):
                clear_table(self.all_tables[x], x)
            self.percentage_data(self.final_dict)
            self.current_field.hide()
            self.file_selector.hide()
            self.next_button.hide()
            self.prev_button.hide()
            self.input_box.hide()
            self.input_box.clear()

    def prev_next(self):
        """
            Update current individual and GUI display based on button click.
            Decrement cur_individual if "previous" button is clicked (unless already at 0).
            Increment cur_individual if "next" button is clicked (unless at end of list).
            Enable/disable "previous" and "next" buttons based on cur_individual.
            Update GUI tables and input fields to display current individual data.
        """
        sender_button = app.sender()
        if sender_button == self.prev_button:
            self.cur_individual = max(0, self.cur_individual - 1)
            self.prev_button.setEnabled(self.cur_individual > 0)
            self.next_button.setEnabled(self.cur_individual < len(self.all_data.keys()) - 1)
        elif sender_button == self.next_button:
            self.cur_individual = min(len(self.all_data.keys()) - 1, self.cur_individual + 1)
            self.prev_button.setEnabled(self.cur_individual > 0)
            self.next_button.setEnabled(self.cur_individual < len(self.all_data.keys()) - 1)
        for x in range(len(self.all_tables)):
            clear_table(self.all_tables[x], x)
            self.input_box.clear()
            data = list(self.all_data.keys())[self.cur_individual]
            self.current_field.setText(str(data))
            index = self.file_selector.findText(str(data))
            self.file_selector.setCurrentIndex(index)
            self.percentage_data(self.all_data[data]["individual"])

    def data(self, all_data):
        """
        Processes the data by extracting the relevant information
        for each trial and classification, sorting it, and storing it
        in a nested dictionary with trial numbers as outer keys and
        classification pairs as inner keys.
        """
        self.prev_button.setEnabled(self.cur_individual > 0)
        self.next_button.setEnabled(self.cur_individual < len(self.all_data.keys()) - 1)

        # Progress bar
        progress = QProgressDialog("Loading files...", None, 0, len(all_data), self)
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setWindowTitle("Loading")
        progress.show()

        # for file_name, data in all_data.items():  # loop over all files and their data
        for num_files, (file_name, data) in enumerate(all_data.items()):
            # Progress loading bar logic
            self.cur_data = {}
            for i in range(1, 9):  # Adding the outermost keys to list (trial number)
                self.cur_data[i] = {}
            for trial_num in range(1, 9):  # all 8 trials **make dynamic**
                trial_idx = data["steady"][:, 3] == trial_num  # get class number
                subset = data["steady"][trial_idx]  # only data for one trial
                classifier = []
                for x in range(len(subset)):
                    classifier.insert(x, subset[x, 2])  # classifier
                    cur_frame = subset[x, 0]  # current start frame
                    if x != 0:  # start after first data entry
                        prev_frame = subset[x - 1, 1]  # previous end frame
                        trial_idx = data["trials"] == trial_num  # get trial data
                        predictions = data["predictions"][trial_idx]  # get predictions data
                        if cur_frame - prev_frame > 1:  # if it does not go directly from one class to other
                            self.cur_data[trial_num][classifier[x - 1], classifier[x]] = list(
                                np.array(predictions[prev_frame:cur_frame]))

            # Sorting the values in the dictionary
            self.cur_data = {trial_num: dict(sorted(values.items(), key=sort_key)) for trial_num, values in
                             self.cur_data.items()}

            self.trial_sorting(self.cur_data, file_name)
            progress.setValue(num_files+1)
            QCoreApplication.processEvents()  # to update the dialog and allow cancellation

    def trial_sorting(self, data, file_name):
        """
            Sorts the data into two dictionaries, `self.final_dict` and `self.individual`, by aggregating and
            storing the data of all individuals in `self.final_dict` and storing the data of the current
            individual in `self.individual`.
            :param file_name: Current file
            :param data: A dictionary containing data for all individuals, sorted by trial and class.
        """
        data_copy = copy.deepcopy(data)
        for trial in data_copy:
            inner_dict = data_copy[trial].copy()
            for inner_key in inner_dict:
                if inner_key in self.final_dict:
                    self.final_dict[inner_key].extend(inner_dict[inner_key])
                else:
                    self.final_dict[inner_key] = inner_dict[inner_key]
        self.percentage_data(self.final_dict)

        cur_individual = {}  # individual person data
        data_copy2 = copy.deepcopy(data)
        for trial in data_copy2:
            inner_dict = data_copy2[trial].copy()
            for inner_key in inner_dict:
                if inner_key in cur_individual:
                    cur_individual[inner_key].extend(inner_dict[inner_key])
                else:
                    cur_individual[inner_key] = inner_dict[inner_key]
        self.all_data[file_name]["individual"] = cur_individual  # store individual data

    def percentage_data(self, data):
        """
            Calculates the percentage of transitions between classes and populates a table
            with the percentage values. It also sets the background color of the cells in
            the table based on the percentage value. If there is no probability for a
            transition, the function sets the cell value to '-' and the total frames to 0.
            :param data: A dictionary containing data for all individuals, sorted by trial and class
        """
        percentages = {}
        for class_trans, numbers in data.items():
            percentages[class_trans] = dict(sorted(count_percentage(numbers).items()))

        for class_trans, transition_frames in percentages.items():  # fill tables
            total_frames = len(data[class_trans[0], class_trans[1]])
            for number, percentage in transition_frames.items():
                table_num = class_trans[0] - 1
                col = number - 1
                row = class_trans[1] - 1
                self.all_tables[table_num].setItem(row, col, QTableWidgetItem(str(percentage) + "%"))
                colour = get_colour(percentage)
                self.all_tables[table_num].item(row, col).setBackground(colour)
                self.all_tables[table_num].item(row, col).setForeground(QtGui.QColor('white'))
                self.all_tables[table_num].setItem(row, 7, QTableWidgetItem(str(total_frames)))

        # if all transitions have no probability
        for x in range(0, 7):
            for row in range(self.all_tables[x].rowCount()):
                for col in range(self.all_tables[x].columnCount()):
                    item = self.all_tables[x].item(row, col)
                    if item is not None and item.text() == "" and item.background().color() != QtGui.QColor(220, 220,
                                                                                                            220):
                        self.all_tables[x].setItem(row, col, QTableWidgetItem("-"))
                        self.all_tables[x].setItem(row, 7, QTableWidgetItem("0"))

    def clear(self):
        self.file_selector.clear()
        self.file_name = None
        self.file_names = None
        self.cur_individual = 0
        self.cur_data = {}
        self.final_dict = {}
        self.all_data = {}

    def screenshot(self):
        screen = QApplication.primaryScreen()
        screenshot = screen.grabWindow(self.winId())

        # save as dataSet_fileName.png
        if self.data_button.isChecked():
            file_name, _ = QFileDialog.getSaveFileName(self, "Save Screenshot", self.data_set.text() + "_" +
                                                       os.path.splitext(self.current_field.text())[0] + ".png",
                                                       "Images (*.png *.xpm *.jpg)")
        else:
            file_name, _ = QFileDialog.getSaveFileName(self, "Save Screenshot", self.data_set.text() + ".png",
                                                       "Images (*.png *.xpm *.jpg)")

        screenshot.save(file_name)

    def export_all_tables(self):
        """
        Exports all tables to an Excel file.
        """
        file_dialog = QFileDialog()
        # save as dataSet_fileName.xlsx
        file_name = self.data_set.text() + ".xlsx"
        file_path, _ = file_dialog.getSaveFileName(self, "Save File", file_name, "Excel Files (*.xlsx)")
        if file_path:
            writer = pd.ExcelWriter(file_path, engine='openpyxl')

            # Write the first sheet as "All Data"
            sheet_name = "All Data"
            self.export_table_to_excel(writer, sheet_name, self.all_tables)

            # Loop through individual data
            for x in range(len(self.all_data.keys())):
                data = list(self.all_data.keys())[x]
                for y in range(len(self.all_tables)):
                    clear_table(self.all_tables[y], y)
                self.percentage_data(self.all_data[data]["individual"])
                sheet_name = data
                self.export_table_to_excel(writer, sheet_name, self.all_tables)

            writer.close()
        for x in range(len(self.all_tables)):
            clear_table(self.all_tables[x], x)
        self.percentage_data(self.final_dict)
        QMessageBox.information(self, "Export Successful", "Tables have been exported to Excel.")


    def export_table_to_excel(self, writer, sheet_name, tables):
        """
        Export a table to an Excel sheet.
        """
        for i, table in enumerate(tables):
            df = self.table_to_dataframe(table)
            df.to_excel(writer, sheet_name=sheet_name, startrow=i * (df.shape[0] + 2), index=False)

            # Get the workbook and corresponding worksheet
            workbook = writer.book
            worksheet = workbook[sheet_name]

            # Apply formatting to row headers (first column)
            for row in range(df.shape[0]):
                cell = worksheet.cell(row=i * (df.shape[0] + 2) + row + 2, column=1)
                cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D3D3D3")
                cell.font = openpyxl.styles.Font(color="000000")

            # Apply formatting to column headers (first row)
            for col in range(df.shape[1]):
                cell = worksheet.cell(row=i * (df.shape[0] + 2) + 1, column=col + 1)
                cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D3D3D3")
                cell.font = openpyxl.styles.Font(color="000000")

            # Apply formatting to specific rows (self-transition rows)
            self_transition_rows = ['NM->NM', 'WF->WF', 'WE->WE', 'WP->WP', 'WS->WS', 'CG->CG', 'HO->HO']
            for row in range(df.shape[0]):
                row_header = df.iloc[row, 0]
                if row_header in self_transition_rows:
                    for col in range(1, df.shape[1]):
                        cell = worksheet.cell(row=i * (df.shape[0] + 2) + row + 2, column=col + 1)
                        cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D3D3D3")
                        cell.font = openpyxl.styles.Font(color="000000")

    def table_to_dataframe(self, table):
        """
        Converts a QTableWidget to a pandas DataFrame.
        :param table: The QTableWidget object
        :type table: QTableWidget
        :return: The corresponding DataFrame
        :rtype: pd.DataFrame
        """
        column_headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
        row_headers = [table.verticalHeaderItem(i).text() for i in range(table.rowCount())]
        data = []
        for row in range(table.rowCount()):
            row_data = []
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)
        df = pd.DataFrame(data, columns=column_headers)
        df.insert(0, '', row_headers)  # Insert row headers as the first column
        return df


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Classifier()
    window.show()
    sys.exit(app.exec())
